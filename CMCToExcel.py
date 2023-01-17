from requests import Request, Session
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import pprint
import json
import requests

#creation du fichier excel
wb = Workbook()
#feuille 1
ws = wb.active
ws.append(["DEVISE","PRIX","MARKETCAP","VOLUME SOUS 24H","TOTAL SUPPLY"])

headers = {'Accepts': 'application/json', 'X-CMC_PRO_API_KEY': key}
key = input("Saisissez votre clé d'API (https://pro.coinmarketcap.com/account) : ")
quotesLatestUrl = 'https://pro-api.coinmarketcap.com/v2/cryptocurrency/quotes/latest'
listedLatestUrl = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'

param = {'start':'1', 'limit':'29'}
response = requests.get(listedLatestUrl, headers=headers, params=param)
jsonResponse = response.json()
#tableau comportant les 30 cryptomonnaies les plus populaires
devises = tuple([d['symbol'] for d in jsonResponse['data']])

#récupération des données et enregistrement dans la feuille excel
for crypto in devises:
	data = {
		'symbol':crypto,
		'convert':'USD'
	}
	try:
		session = Session()
		session.headers.update(headers)
		response = session.get(quotesLatestUrl, params = data)
		price = round((json.loads(response.text)['data'][crypto][0]['quote']['USD']['price']),2)
		marketcap = json.loads(response.text)['data'][crypto][0]['quote']['USD']['market_cap']
		volume24h = json.loads(response.text)['data'][crypto][0]['quote']['USD']['volume_24h']
		totalSupply = json.loads(response.text)['data'][crypto][0]['total_supply']
	except KeyError:
		price = "Erreur : taux dépassé"
		marketcap = "-"
		volume24h = "-"
		totalSupply = "-"
	# ajout d'une ligne contenant la devise et le prix correspondant
	ws.append(["$"+ crypto, price, marketcap,volume24h,totalSupply])

#personnalisation de la feuille excel
#police
for cell in ws[1]:
    cell.font = Font(name='Arial', bold=True)

#taille automatique des colonnes
for col in ws.columns:
    max_length = 0
    column = col[0].column
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

#on colore la colonne contenant les devises
for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, max_row=ws.max_row):
    for cell in row:
        cell.fill = PatternFill(start_color="B0F2B6", end_color="B0F2B6", fill_type = "solid")

#entêtes
for cell in ws[1]:
    cell.fill = PatternFill(start_color="096A09", end_color="096A09", fill_type = "solid")
for cell in ws[1]:
    cell.font = Font(color="FFFFFF")
for row in ws.iter_rows():
    for cell in row:
        cell.font = Font(size=14)



wb.save('data.xlsx')
